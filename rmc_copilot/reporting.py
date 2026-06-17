from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def ajustar_largura_colunas(ws, largura_maxima=60):
    """
    Ajusta largura das colunas automaticamente.
    """
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(value))
            except Exception:
                pass

        adjusted_width = min(max_length + 2, largura_maxima)
        ws.column_dimensions[col_letter].width = adjusted_width


def aplicar_estilo_tabela(ws):
    """
    Aplica estilo simples no cabeçalho da planilha.
    """
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="D9EAF7")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def criar_aba_dashboard(wb):
    """
    Cria aba DASHBOARD.
    """
    if "DASHBOARD" in wb.sheetnames:
        del wb["DASHBOARD"]

    ws = wb.create_sheet("DASHBOARD", 0)

    ws["A1"] = "RMC Copilot v0.5 — Dashboard Executivo"
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")

    ws["A3"] = "Este dashboard consolida prioridades operacionais, risco futuro, clusters críticos e VMs com maior score."
    ws["A3"].font = Font(size=11)

    ws["A5"] = "Indicadores principais"
    ws["A5"].font = Font(size=14, bold=True, color="1F4E78")

    return ws


def escrever_kpis_dashboard(ws, df_analise_v4):
    """
    Escreve KPIs principais no Dashboard.
    """
    total = len(df_analise_v4)

    prioridade = df_analise_v4["prioridade_final"].value_counts()
    acao = df_analise_v4["acao_final"].value_counts()
    risco_futuro = df_analise_v4["criticidade_futura"].value_counts()

    p0 = int(prioridade.get("P0_ACAO_IMEDIATA", 0))
    p1 = int(prioridade.get("P1_ALTA", 0))
    p2 = int(prioridade.get("P2_MEDIA", 0))
    p3 = int(prioridade.get("P3_BAIXA", 0))
    p4 = int(prioridade.get("P4_MONITORAR", 0))

    risco_30 = int(risco_futuro.get("RISCO_FUTURO_30D", 0))
    risco_60 = int(risco_futuro.get("RISCO_FUTURO_60D", 0))
    risco_90 = int(risco_futuro.get("RISCO_FUTURO_90D", 0))

    kpis = [
        ("Total de VMs analisadas", total),
        ("P0 — Ação imediata", p0),
        ("P1 — Alta prioridade", p1),
        ("P2 — Média prioridade", p2),
        ("P3 — Baixa prioridade", p3),
        ("P4 — Monitorar", p4),
        ("Risco futuro 30 dias", risco_30),
        ("Risco futuro 60 dias", risco_60),
        ("Risco futuro 90 dias", risco_90),
    ]

    start_row = 6

    ws["A6"] = "Indicador"
    ws["B6"] = "Valor"

    for idx, (nome, valor) in enumerate(kpis, start=start_row + 1):
        ws[f"A{idx}"] = nome
        ws[f"B{idx}"] = valor

    for cell in ws[6]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)

    for row in range(7, 7 + len(kpis)):
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].alignment = Alignment(horizontal="center")


def criar_grafico_barra(
    ws,
    titulo,
    min_col,
    max_col,
    min_row,
    max_row,
    posicao,
    altura=8,
    largura=16
):
    """
    Cria gráfico de barras.
    """
    chart = BarChart()
    chart.title = titulo
    chart.y_axis.title = "Quantidade"
    chart.x_axis.title = "Categoria"

    data = Reference(ws, min_col=max_col, max_col=max_col, min_row=min_row, max_row=max_row)
    cats = Reference(ws, min_col=min_col, max_col=min_col, min_row=min_row + 1, max_row=max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    chart.height = altura
    chart.width = largura

    ws.add_chart(chart, posicao)


def criar_grafico_pizza(
    ws,
    titulo,
    min_col,
    max_col,
    min_row,
    max_row,
    posicao,
    altura=8,
    largura=12
):
    """
    Cria gráfico de pizza.
    """
    chart = PieChart()
    chart.title = titulo

    data = Reference(ws, min_col=max_col, max_col=max_col, min_row=min_row, max_row=max_row)
    cats = Reference(ws, min_col=min_col, max_col=min_col, min_row=min_row + 1, max_row=max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    chart.height = altura
    chart.width = largura

    ws.add_chart(chart, posicao)


def criar_tabelas_resumo_dashboard(wb, df_analise_v4, resumo_cluster_v04):
    """
    Cria tabelas auxiliares na aba DASHBOARD para alimentar gráficos.
    """
    ws = wb["DASHBOARD"]

    prioridade = (
        df_analise_v4["prioridade_final"]
        .value_counts()
        .reset_index()
    )
    prioridade.columns = ["prioridade_final", "quantidade"]

    acao = (
        df_analise_v4["acao_final"]
        .value_counts()
        .reset_index()
    )
    acao.columns = ["acao_final", "quantidade"]

    risco_futuro = (
        df_analise_v4["criticidade_futura"]
        .value_counts()
        .reset_index()
    )
    risco_futuro.columns = ["criticidade_futura", "quantidade"]

    top_clusters = resumo_cluster_v04[
        [
            "cluster",
            "p0_acao_imediata",
            "p1_alta",
            "vms_prioritarias",
            "pct_vms_prioritarias",
            "score_max",
        ]
    ].head(10).copy()

    # Prioridade
    start_col = 4
    start_row = 6

    ws.cell(start_row, start_col, "prioridade_final")
    ws.cell(start_row, start_col + 1, "quantidade")

    for i, row in prioridade.iterrows():
        ws.cell(start_row + i + 1, start_col, row["prioridade_final"])
        ws.cell(start_row + i + 1, start_col + 1, int(row["quantidade"]))

    # Ações
    start_col = 7
    start_row = 6

    ws.cell(start_row, start_col, "acao_final")
    ws.cell(start_row, start_col + 1, "quantidade")

    for i, row in acao.iterrows():
        ws.cell(start_row + i + 1, start_col, row["acao_final"])
        ws.cell(start_row + i + 1, start_col + 1, int(row["quantidade"]))

    # Risco futuro
    start_col = 10
    start_row = 6

    ws.cell(start_row, start_col, "criticidade_futura")
    ws.cell(start_row, start_col + 1, "quantidade")

    for i, row in risco_futuro.iterrows():
        ws.cell(start_row + i + 1, start_col, row["criticidade_futura"])
        ws.cell(start_row + i + 1, start_col + 1, int(row["quantidade"]))

    # Top clusters
    start_col = 13
    start_row = 6

    for j, col in enumerate(top_clusters.columns, start=start_col):
        ws.cell(start_row, j, col)

    for i, (_, row) in enumerate(top_clusters.iterrows(), start=start_row + 1):
        for j, col in enumerate(top_clusters.columns, start=start_col):
            ws.cell(i, j, row[col])

    return {
        "prioridade_rows": len(prioridade) + 1,
        "acao_rows": len(acao) + 1,
        "risco_futuro_rows": len(risco_futuro) + 1,
        "top_clusters_rows": len(top_clusters) + 1,
    }


def criar_graficos_dashboard(wb, meta):
    """
    Cria gráficos na aba Dashboard.
    """
    ws = wb["DASHBOARD"]

    # Prioridade: D/E
    criar_grafico_pizza(
        ws=ws,
        titulo="Distribuição por Prioridade Final",
        min_col=4,
        max_col=5,
        min_row=6,
        max_row=6 + meta["prioridade_rows"] - 1,
        posicao="A18",
    )

    # Ações: G/H
    criar_grafico_barra(
        ws=ws,
        titulo="Distribuição por Ação Final",
        min_col=7,
        max_col=8,
        min_row=6,
        max_row=6 + meta["acao_rows"] - 1,
        posicao="I18",
        altura=9,
        largura=18,
    )

    # Risco futuro: J/K
    criar_grafico_barra(
        ws=ws,
        titulo="Risco Futuro 30/60/90 dias",
        min_col=10,
        max_col=11,
        min_row=6,
        max_row=6 + meta["risco_futuro_rows"] - 1,
        posicao="A35",
        altura=8,
        largura=16,
    )

    # Top clusters: M/O
    criar_grafico_barra(
        ws=ws,
        titulo="Top Clusters por VMs Prioritárias",
        min_col=13,
        max_col=15,
        min_row=6,
        max_row=6 + meta["top_clusters_rows"] - 1,
        posicao="I35",
        altura=8,
        largura=18,
    )


def gerar_resumo_executivo_df(resumo_executivo: str) -> pd.DataFrame:
    """
    Transforma texto do resumo executivo em DataFrame.
    """
    return pd.DataFrame({"resumo_executivo": resumo_executivo.split("\n")})


def gerar_relatorio_excel_v05(
    caminho_saida,
    df_analise_v4,
    resumo_cluster_v04,
    resumo_executivo_v04
):
    """
    Gera Excel executivo v0.5 com dashboard e gráficos.
    """

    caminho_saida = Path(caminho_saida)
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    top_vms = df_analise_v4[
        [
            "cluster",
            "vm",
            "categoria_vm",
            "status_geral",
            "risco_futuro_90d",
            "criticidade_futura",
            "score_prioridade",
            "prioridade_final",
            "acao_final",
            "cpu_p95_pct",
            "mem_p95_pct",
            "disk_p95_pct",
            "cpu_forecast_90d",
            "mem_forecast_90d",
            "disk_forecast_90d",
            "recomendacao_final",
        ]
    ].head(100).copy()

    df_p0 = df_analise_v4[df_analise_v4["prioridade_final"] == "P0_ACAO_IMEDIATA"].copy()
    df_p1 = df_analise_v4[df_analise_v4["prioridade_final"] == "P1_ALTA"].copy()

    df_risco_futuro = df_analise_v4[
        df_analise_v4["risco_futuro_90d"] != "SEM_RISCO_90D"
    ].copy()

    resumo_prioridade = df_analise_v4["prioridade_final"].value_counts().reset_index()
    resumo_prioridade.columns = ["prioridade_final", "quantidade"]

    resumo_acao = df_analise_v4["acao_final"].value_counts().reset_index()
    resumo_acao.columns = ["acao_final", "quantidade"]

    resumo_criticidade_futura = df_analise_v4["criticidade_futura"].value_counts().reset_index()
    resumo_criticidade_futura.columns = ["criticidade_futura", "quantidade"]

    resumo_categoria = df_analise_v4["categoria_vm"].value_counts().reset_index()
    resumo_categoria.columns = ["categoria_vm", "quantidade"]

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        gerar_resumo_executivo_df(resumo_executivo_v04).to_excel(
            writer,
            sheet_name="RESUMO_EXECUTIVO",
            index=False
        )

        resumo_cluster_v04.to_excel(
            writer,
            sheet_name="TOP_CLUSTERS",
            index=False
        )

        top_vms.to_excel(
            writer,
            sheet_name="TOP_VMS",
            index=False
        )

        df_p0.to_excel(
            writer,
            sheet_name="P0_ACAO_IMEDIATA",
            index=False
        )

        df_p1.to_excel(
            writer,
            sheet_name="P1_ALTA",
            index=False
        )

        df_risco_futuro.to_excel(
            writer,
            sheet_name="RISCO_FUTURO",
            index=False
        )

        resumo_prioridade.to_excel(
            writer,
            sheet_name="RESUMO_PRIORIDADE",
            index=False
        )

        resumo_acao.to_excel(
            writer,
            sheet_name="RESUMO_ACAO_FINAL",
            index=False
        )

        resumo_criticidade_futura.to_excel(
            writer,
            sheet_name="RESUMO_RISCO_FUTURO",
            index=False
        )

        resumo_categoria.to_excel(
            writer,
            sheet_name="RESUMO_CATEGORIA_VM",
            index=False
        )

        df_analise_v4.to_excel(
            writer,
            sheet_name="ANALISE_COMPLETA",
            index=False
        )

    wb = load_workbook(caminho_saida)

    criar_aba_dashboard(wb)
    escrever_kpis_dashboard(wb["DASHBOARD"], df_analise_v4)
    meta = criar_tabelas_resumo_dashboard(wb, df_analise_v4, resumo_cluster_v04)
    criar_graficos_dashboard(wb, meta)

    for ws in wb.worksheets:
        aplicar_estilo_tabela(ws)
        ajustar_largura_colunas(ws)

    wb.save(caminho_saida)

    return caminho_saida